from pathlib import Path
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

def excel(cuit):
    # Ruta a la carpeta de Descargas
    downloads_path = Path.home() / "Downloads"
    
    buy_file(downloads_path, cuit) 
    sell_file(downloads_path, cuit)
    

def sell_file(downloads_path, cuit):
    sell_file = downloads_path / f"Comprobantes de Ventas - CUIT {cuit}.xlsx"
    if not sell_file.exists():
        print(f"El archivo {sell_file} no existe. Hubo un error al descargar el archivo de AFIP.")
        return
    
    # Abrir directamente con openpyxl
    wb = load_workbook(sell_file)
    ws = wb.active 
    
    sell_column_dimensions(ws)
    check_sell(ws)
    autosum_last_row(ws)
    
    
    # Guardar en una copia nueva
    final_file = downloads_path / f"Ventas_modificadas_{cuit}.xlsx"
    wb.save(final_file)

    print(f"Archivo modificado guardado como: {final_file}")
    
def buy_file(downloads_path, cuit):
    compras_file = downloads_path / f"Comprobantes de Compras - CUIT {cuit}.xlsx"
    if not compras_file.exists():
        print(f"El archivo {compras_file} no existe. Hubo un error al descargar el archivo de AFIP.")
        return

    # Abrir directamente con openpyxl
    wb = load_workbook(compras_file)
    ws = wb.active  

    buy_column_dimensions(ws)
    check_buy(ws)
    autosum_last_row(ws)
    
    # Guardar en una copia nueva
    final_file = downloads_path / f"Compras_modificado_{cuit}.xlsx"
    wb.save(final_file)

    print(f"Archivo modificado guardado como: {final_file}")
    
def sell_column_dimensions(ws):
    # Cambiar ancho de columnas
    ws.column_dimensions["A"].width = 10.2857
    ws.column_dimensions["B"].width = 15.71428571428571
    ws.column_dimensions["C"].width = 4.285714285714286
    ws.column_dimensions["D"].width = 5.714285714285714
    ws.column_dimensions["E"].width = 4.285714285714286
    ws.column_dimensions["F"].width = 5
    ws.column_dimensions["G"].width = 12.85714285714286
    ws.column_dimensions["H"].width = 30.71428571428571
    ws.column_dimensions["I"].width = 3.571428571428571
    ws.column_dimensions["J"].width = 2.857142857142857
    ws.column_dimensions["K"].width = 13.57142857142857
    ws.column_dimensions["L"].width = 11.42857142857143
    
def buy_column_dimensions(ws):
    # Cambiar ancho de columnas
    ws.column_dimensions["A"].width = 10.714
    ws.column_dimensions["B"].width = 12.5715
    ws.column_dimensions["C"].width = 5.714
    ws.column_dimensions["D"].width = 8.571
    ws.column_dimensions["E"].width = 4.285
    ws.column_dimensions["F"].width = 5
    ws.column_dimensions["G"].width = 12.571
    ws.column_dimensions["H"].width = 26.857
    ws.column_dimensions["I"].width = 4.285
    ws.column_dimensions["J"].width = 5
    ws.column_dimensions["K"].width = 12.857
    ws.column_dimensions["L"].width = 12.143
        
def autosum_last_row(ws):
    # Buscar la última fila con datos numéricos desde la fila 3 en adelante
    last_data_row = 3
    for row in range(3, ws.max_row + 1):
        # Verificamos si hay datos numéricos en alguna celda entre K y O
        has_data = False
        for col in range(11, 16):  # Columnas K (11) a O (15)
            cell = ws.cell(row=row, column=col)
            if isinstance(cell.value, (int, float)):
                has_data = True
                break
        if has_data:
            last_data_row = row

    # Fila donde escribir las fórmulas: dos filas debajo de la última con datos
    formula_row = last_data_row + 2

    # Insertar las fórmulas en las columnas K a O
    for col in range(11, 16):
        col_letter = get_column_letter(col)
        formula = f"=SUM({col_letter}3:{col_letter}{last_data_row})"
        ws[f"{col_letter}{formula_row}"] = formula 

def check_sell(ws):
    columnas_objetivo = ["K", "L", "M", "N", "O"]

    for fila in range(3, ws.max_row + 1):
        tipo_factura = ws[f"B{fila}"].value  # Columna B

        if isinstance(tipo_factura, str):
            tipo_factura_normalizado = tipo_factura.lower().strip()

            # Si contiene "nota de crédito" y es de tipo A, B o C
            if "nota de crédito a" in tipo_factura_normalizado or \
               "nota de crédito b" in tipo_factura_normalizado or \
               "nota de crédito c" in tipo_factura_normalizado:
                
                for col in columnas_objetivo:
                    celda = ws[f"{col}{fila}"]
                    if isinstance(celda.value, (int, float)):
                        celda.value = -abs(celda.value)

def check_buy(ws):
    # 1️⃣ Encontrar la última fila con datos
    ultima_fila = ws.max_row
    while ultima_fila > 0 and all(cell.value is None for cell in ws[ultima_fila]):
        ultima_fila -= 1

    # 2️⃣ Buscar filas con "Factura B" o "Factura C" en la columna B
    filas_a_mover = []
    for row in range(1, ultima_fila + 1):
        valor = str(ws.cell(row=row, column=2).value).strip() if ws.cell(row=row, column=2).value else ""
        if "Factura B" in valor or "Factura C" in valor:
            filas_a_mover.append(row)

    # Si no hay coincidencias, salir
    if not filas_a_mover:
        print("No se encontraron Facturas B o C.")
        return

    # 3️⃣ Copiar datos de las filas encontradas
    datos_filas = []
    for fila in filas_a_mover:
        datos_filas.append([ws.cell(row=fila, column=col).value for col in range(1, ws.max_column + 1)])

    # 4️⃣ Eliminar filas desde abajo hacia arriba (para no alterar índices)
    for fila in reversed(filas_a_mover):
        ws.delete_rows(fila, 1)

    # 5️⃣ Insertar las filas 4 líneas debajo de la nueva última fila
    nueva_ultima = ws.max_row
    posicion_insercion = nueva_ultima + 4

    for datos in datos_filas:
        ws.insert_rows(posicion_insercion)
        for col, valor in enumerate(datos, start=1):
            ws.cell(row=posicion_insercion, column=col).value = valor
        posicion_insercion += 1